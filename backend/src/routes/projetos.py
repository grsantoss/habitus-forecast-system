from flask import Blueprint, request, jsonify, send_file
from datetime import datetime, timedelta
from sqlalchemy import func, extract
from src.models.user import db, Projeto, Cenario, LogSistema, ArquivoUpload, LancamentoFinanceiro, CategoriaFinanceira, HistoricoCenario, User, Relatorio
from src.auth import token_required, admin_required
from io import BytesIO
import os

projetos_bp = Blueprint('projetos', __name__)

@projetos_bp.route('/projetos', methods=['GET'])
@token_required
def listar_projetos(current_user):
    """Lista projetos do usuário atual ou todos (se admin)"""
    try:
        if current_user.role == 'admin':
            projetos = Projeto.query.all()
        else:
            projetos = Projeto.query.filter_by(usuario_id=current_user.id).all()
        
        projetos_data = []
        for projeto in projetos:
            d = projeto.to_dict()
            # Anexar nome_arquivo do upload mais recente (se houver)
            try:
                upload_recente = ArquivoUpload.query.filter_by(projeto_id=projeto.id)\
                    .order_by(ArquivoUpload.uploaded_at.desc()).first()
                d['nome_arquivo'] = upload_recente.nome_original if upload_recente else None
            except Exception:
                d['nome_arquivo'] = None
            projetos_data.append(d)

        return jsonify({
            'projetos': projetos_data
        })
        
    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/projetos', methods=['POST'])
@token_required
def criar_projeto(current_user):
    """Cria um novo projeto"""
    try:
        data = request.get_json()
        
        if not data or not data.get('nome_cliente') or not data.get('data_base_estudo'):
            return jsonify({'message': 'Nome do cliente e data base são obrigatórios'}), 400
        
        # Criar projeto
        projeto = Projeto(
            usuario_id=current_user.id,
            nome_cliente=data.get('nome_cliente'),
            data_base_estudo=datetime.strptime(data.get('data_base_estudo'), '%Y-%m-%d').date(),
            saldo_inicial_caixa=data.get('saldo_inicial_caixa', 0)
        )
        
        db.session.add(projeto)
        db.session.flush()  # Para obter o ID do projeto
        
        # Criar cenário padrão "Realista"
        cenario_padrao = Cenario(
            projeto_id=projeto.id,
            nome='Realista',
            descricao='Cenário padrão baseado em projeções realistas',
            is_active=True
        )
        
        db.session.add(cenario_padrao)
        db.session.commit()
        
        # Log da criação
        log = LogSistema(
            usuario_id=current_user.id,
            acao='PROJECT_CREATED',
            detalhes={
                'projeto_id': projeto.id,
                'nome_cliente': projeto.nome_cliente
            }
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'message': 'Projeto criado com sucesso',
            'projeto': projeto.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/projetos/<int:projeto_id>', methods=['GET'])
@token_required
def obter_projeto(current_user, projeto_id):
    """Obtém detalhes de um projeto específico"""
    try:
        projeto = Projeto.query.get_or_404(projeto_id)
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        # Incluir cenários do projeto
        projeto_dict = projeto.to_dict()
        projeto_dict['cenarios'] = [cenario.to_dict() for cenario in projeto.cenarios]
        
        return jsonify({'projeto': projeto_dict})
        
    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/projetos/<int:projeto_id>', methods=['PUT'])
@token_required
def atualizar_projeto(current_user, projeto_id):
    """Atualiza um projeto"""
    try:
        projeto = Projeto.query.get_or_404(projeto_id)
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        data = request.get_json()
        
        # Atualizar campos
        if data.get('nome_cliente'):
            projeto.nome_cliente = data.get('nome_cliente')
        if data.get('data_base_estudo'):
            projeto.data_base_estudo = datetime.strptime(data.get('data_base_estudo'), '%Y-%m-%d').date()
        if data.get('saldo_inicial_caixa') is not None:
            projeto.saldo_inicial_caixa = data.get('saldo_inicial_caixa')
        
        projeto.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        # Log da atualização
        log = LogSistema(
            usuario_id=current_user.id,
            acao='PROJECT_UPDATED',
            detalhes={
                'projeto_id': projeto.id,
                'nome_cliente': projeto.nome_cliente
            }
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'message': 'Projeto atualizado com sucesso',
            'projeto': projeto.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/projetos/<int:projeto_id>', methods=['DELETE'])
@token_required
def deletar_projeto(current_user, projeto_id):
    """Deleta um projeto"""
    try:
        projeto = Projeto.query.get_or_404(projeto_id)
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        nome_cliente = projeto.nome_cliente
        
        # Log antes de deletar
        log = LogSistema(
            usuario_id=current_user.id,
            acao='PROJECT_DELETED',
            detalhes={
                'projeto_id': projeto.id,
                'nome_cliente': nome_cliente
            }
        )
        db.session.add(log)
        
        db.session.delete(projeto)
        db.session.commit()
        
        return jsonify({'message': 'Projeto deletado com sucesso'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/cenarios', methods=['GET'])
@token_required
def listar_cenarios(current_user):
    """Lista todos os cenários do usuário atual ou todos (se admin)"""
    try:
        # Buscar projetos do usuário
        if current_user.role == 'admin':
            projetos = Projeto.query.all()
        else:
            projetos = Projeto.query.filter_by(usuario_id=current_user.id).all()
        
        projeto_ids = [p.id for p in projetos]
        
        if not projeto_ids:
            return jsonify({'cenarios': []}), 200
        
        # Buscar todos os cenários dos projetos do usuário
        cenarios = Cenario.query.filter(Cenario.projeto_id.in_(projeto_ids)).order_by(Cenario.created_at.desc()).all()
        
        # Buscar informações do projeto e arquivo para cada cenário
        cenarios_data = []
        for cenario in cenarios:
            projeto = Projeto.query.get(cenario.projeto_id)
            upload_recente = ArquivoUpload.query.filter_by(projeto_id=cenario.projeto_id)\
                .order_by(ArquivoUpload.uploaded_at.desc()).first()
            
            cenario_dict = cenario.to_dict()
            cenario_dict['projeto_nome'] = projeto.nome_cliente if projeto else 'Projeto Desconhecido'
            cenario_dict['arquivo_nome'] = upload_recente.nome_original if upload_recente else None
            cenario_dict['arquivo_data'] = upload_recente.uploaded_at.isoformat() if upload_recente and upload_recente.uploaded_at else None
            
            # Calcular estatísticas básicas (receita total, etc)
            # Isso pode ser otimizado depois, mas por enquanto vamos retornar dados básicos
            cenarios_data.append(cenario_dict)
        
        return jsonify({'cenarios': cenarios_data}), 200
        
    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/cenarios/comparar', methods=['POST'])
@token_required
def comparar_cenarios(current_user):
    """Compara múltiplos cenários e retorna estatísticas comparativas"""
    try:
        data = request.get_json()
        
        if not data or not data.get('cenario_ids'):
            return jsonify({'message': 'Lista de IDs de cenários é obrigatória'}), 400
        
        cenario_ids = data.get('cenario_ids')
        
        if not isinstance(cenario_ids, list) or len(cenario_ids) < 2:
            return jsonify({'message': 'É necessário pelo menos 2 cenários para comparar'}), 400
        
        if len(cenario_ids) > 5:
            return jsonify({'message': 'Máximo de 5 cenários por comparação'}), 400
        
        # Buscar cenários e verificar permissões
        cenarios = []
        for cenario_id in cenario_ids:
            cenario = Cenario.query.get(cenario_id)
            if not cenario:
                return jsonify({'message': f'Cenário {cenario_id} não encontrado'}), 404
            
            projeto = Projeto.query.get(cenario.projeto_id)
            if not projeto:
                return jsonify({'message': f'Projeto do cenário {cenario_id} não encontrado'}), 404
            
            # Verificar permissão
            if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
                return jsonify({'message': f'Acesso negado ao cenário {cenario_id}'}), 403
            
            cenarios.append({
                'cenario': cenario,
                'projeto': projeto
            })
        
        # Buscar análises de cada cenário
        comparacao = []
        for item in cenarios:
            cenario = item['cenario']
            projeto = item['projeto']
            
            # Buscar lançamentos
            lancamentos = LancamentoFinanceiro.query.filter_by(cenario_id=cenario.id).all()
            
            # Calcular estatísticas
            total_entradas = sum(float(l.valor) for l in lancamentos if l.tipo == 'ENTRADA')
            total_saidas = sum(float(l.valor) for l in lancamentos if l.tipo == 'SAIDA')
            saldo_liquido = total_entradas - total_saidas
            total_lancamentos = len(lancamentos)
            
            # Buscar arquivo relacionado
            upload_recente = ArquivoUpload.query.filter_by(projeto_id=projeto.id)\
                .order_by(ArquivoUpload.uploaded_at.desc()).first()
            
            # Agrupar por período (mensal) para gráficos
            dados_agrupados = {}
            for lancamento in lancamentos:
                data = lancamento.data_competencia
                chave = f"{data.year}-{data.month:02d}"
                meses_pt = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
                periodo_label = f"{meses_pt[data.month - 1]}/{data.year}"
                
                if chave not in dados_agrupados:
                    dados_agrupados[chave] = {
                        'periodo': periodo_label,
                        'entradas': 0,
                        'saidas': 0,
                        'saldo_liquido': 0
                    }
                
                valor = float(lancamento.valor)
                if lancamento.tipo == 'ENTRADA':
                    dados_agrupados[chave]['entradas'] += valor
                else:
                    dados_agrupados[chave]['saidas'] += valor
                
                dados_agrupados[chave]['saldo_liquido'] = dados_agrupados[chave]['entradas'] - dados_agrupados[chave]['saidas']
            
            fluxo_caixa = [
                {
                    'periodo': dados['periodo'],
                    'entradas': dados['entradas'],
                    'saidas': dados['saidas'],
                    'saldo_liquido': dados['saldo_liquido']
                }
                for chave, dados in sorted(dados_agrupados.items())
            ]
            
            comparacao.append({
                'cenario_id': cenario.id,
                'cenario_nome': cenario.nome,
                'projeto_nome': projeto.nome_cliente,
                'is_active': cenario.is_active,
                'estatisticas': {
                    'total_entradas': total_entradas,
                    'total_saidas': total_saidas,
                    'saldo_liquido': saldo_liquido,
                    'total_lancamentos': total_lancamentos
                },
                'fluxo_caixa': fluxo_caixa,
                'arquivo_nome': upload_recente.nome_original if upload_recente else None
            })
        
        # Calcular diferenças percentuais (usando o primeiro cenário como base)
        if len(comparacao) >= 2:
            base = comparacao[0]
            base_saldo = base['estatisticas']['saldo_liquido']
            base_entradas = base['estatisticas']['total_entradas']
            base_saidas = base['estatisticas']['total_saidas']
            
            for item in comparacao[1:]:
                item['diferencas_percentuais'] = {
                    'saldo_liquido': (
                        ((item['estatisticas']['saldo_liquido'] - base_saldo) / base_saldo * 100)
                        if base_saldo != 0 else 0
                    ),
                    'entradas': (
                        ((item['estatisticas']['total_entradas'] - base_entradas) / base_entradas * 100)
                        if base_entradas != 0 else 0
                    ),
                    'saidas': (
                        ((item['estatisticas']['total_saidas'] - base_saidas) / base_saidas * 100)
                        if base_saidas != 0 else 0
                    )
                }
            
            # Adicionar diferença zero para o primeiro (base)
            comparacao[0]['diferencas_percentuais'] = {
                'saldo_liquido': 0,
                'entradas': 0,
                'saidas': 0
            }
        
        return jsonify({
            'comparacao': comparacao,
            'total_cenarios': len(comparacao)
        }), 200
        
    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/projetos/<int:projeto_id>/cenarios', methods=['POST'])
@token_required
def criar_cenario(current_user, projeto_id):
    """Cria um novo cenário para um projeto"""
    try:
        projeto = Projeto.query.get_or_404(projeto_id)
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        data = request.get_json()
        
        if not data or not data.get('nome'):
            return jsonify({'message': 'Nome do cenário é obrigatório'}), 400
        
        # Criar cenário
        cenario = Cenario(
            projeto_id=projeto_id,
            nome=data.get('nome'),
            descricao=data.get('descricao', ''),
            is_active=data.get('is_active', False)
        )
        
        db.session.add(cenario)
        db.session.commit()
        
        # Log da criação
        log = LogSistema(
            usuario_id=current_user.id,
            acao='SCENARIO_CREATED',
            detalhes={
                'projeto_id': projeto_id,
                'cenario_id': cenario.id,
                'nome_cenario': cenario.nome
            }
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'message': 'Cenário criado com sucesso',
            'cenario': cenario.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/cenarios/<int:cenario_id>', methods=['PUT'])
@token_required
def atualizar_cenario(current_user, cenario_id):
    """Atualiza um cenário (status, nome, descrição)"""
    try:
        cenario = Cenario.query.get_or_404(cenario_id)
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        data = request.get_json()
        
        if data.get('is_active') is not None:
            cenario.is_active = data.get('is_active')
        if data.get('nome'):
            cenario.nome = data.get('nome')
        if data.get('descricao') is not None:
            cenario.descricao = data.get('descricao')
        
        db.session.commit()
        
        # Log da atualização
        log = LogSistema(
            usuario_id=current_user.id,
            acao='SCENARIO_UPDATED',
            detalhes={
                'cenario_id': cenario_id,
                'nome_cenario': cenario.nome,
                'is_active': cenario.is_active
            }
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'message': 'Cenário atualizado com sucesso',
            'cenario': cenario.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/cenarios/<int:cenario_id>', methods=['DELETE'])
@token_required
def deletar_cenario(current_user, cenario_id):
    """Deleta um cenário"""
    try:
        cenario = Cenario.query.get_or_404(cenario_id)
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        nome_cenario = cenario.nome
        projeto_id = cenario.projeto_id
        
        # Log antes de deletar
        log = LogSistema(
            usuario_id=current_user.id,
            acao='SCENARIO_DELETED',
            detalhes={
                'cenario_id': cenario_id,
                'nome_cenario': nome_cenario,
                'projeto_id': projeto_id
            }
        )
        db.session.add(log)
        
        db.session.delete(cenario)
        db.session.commit()
        
        return jsonify({'message': f'Cenário {nome_cenario} deletado com sucesso'}), 200
                
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/cenarios/<int:cenario_id>/analise', methods=['GET'])
@token_required
def analisar_cenario(current_user, cenario_id):
    """Retorna análise detalhada de um cenário"""
    try:
        cenario = Cenario.query.get_or_404(cenario_id)
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        # Buscar lançamentos do cenário
        lancamentos = LancamentoFinanceiro.query.filter_by(cenario_id=cenario_id).all()
        
        # Estatísticas básicas
        total_entradas = sum(float(l.valor) for l in lancamentos if l.tipo == 'ENTRADA')
        total_saidas = sum(float(l.valor) for l in lancamentos if l.tipo == 'SAIDA')
        saldo_liquido = total_entradas - total_saidas
        
        # Contagem de lançamentos
        total_lancamentos = len(lancamentos)
        lancamentos_entrada = len([l for l in lancamentos if l.tipo == 'ENTRADA'])
        lancamentos_saida = len([l for l in lancamentos if l.tipo == 'SAIDA'])
        lancamentos_projetados = len([l for l in lancamentos if l.origem == 'PROJETADO'])
        lancamentos_realizados = len([l for l in lancamentos if l.origem == 'REALIZADO'])
        
        # Buscar arquivo relacionado
        upload_recente = ArquivoUpload.query.filter_by(projeto_id=projeto.id)\
            .order_by(ArquivoUpload.uploaded_at.desc()).first()
        
        # Estatísticas por categoria
        lancamentos_por_categoria = db.session.query(
            CategoriaFinanceira.nome,
            func.sum(LancamentoFinanceiro.valor).label('total'),
            func.count(LancamentoFinanceiro.id).label('quantidade')
        ).join(
            LancamentoFinanceiro, CategoriaFinanceira.id == LancamentoFinanceiro.categoria_id
        ).filter(
            LancamentoFinanceiro.cenario_id == cenario_id
        ).group_by(
            CategoriaFinanceira.nome
        ).all()
        
        categorias_data = [
            {
                'nome': row.nome,
                'total': float(row.total) if row.total else 0,
                'quantidade': row.quantidade
            }
            for row in lancamentos_por_categoria
        ]
        
        # Período de análise (primeira e última data)
        datas = [l.data_competencia for l in lancamentos if l.data_competencia]
        periodo_inicio = min(datas).isoformat() if datas else None
        periodo_fim = max(datas).isoformat() if datas else None
        
        analise = {
            'cenario': cenario.to_dict(),
            'projeto': {
                'id': projeto.id,
                'nome_cliente': projeto.nome_cliente,
                'data_base_estudo': projeto.data_base_estudo.isoformat() if projeto.data_base_estudo else None,
                'saldo_inicial_caixa': float(projeto.saldo_inicial_caixa) if projeto.saldo_inicial_caixa else 0
            },
            'arquivo': {
                'nome_original': upload_recente.nome_original if upload_recente else None,
                'uploaded_at': upload_recente.uploaded_at.isoformat() if upload_recente and upload_recente.uploaded_at else None
            },
            'estatisticas': {
                'total_entradas': total_entradas,
                'total_saidas': total_saidas,
                'saldo_liquido': saldo_liquido,
                'total_lancamentos': total_lancamentos,
                'lancamentos_entrada': lancamentos_entrada,
                'lancamentos_saida': lancamentos_saida,
                'lancamentos_projetados': lancamentos_projetados,
                'lancamentos_realizados': lancamentos_realizados,
                'periodo_inicio': periodo_inicio,
                'periodo_fim': periodo_fim
            },
            'categorias': categorias_data
        }
        
        return jsonify(analise), 200
        
    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/cenarios/<int:cenario_id>/lancamentos', methods=['POST'])
@token_required
def criar_lancamento(current_user, cenario_id):
    """Cria um novo lançamento financeiro em um cenário"""
    try:
        cenario = Cenario.query.get_or_404(cenario_id)
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        # Validar que o cenário está ativo para edição
        if not cenario.is_active:
            return jsonify({'message': 'Não é possível editar lançamentos em cenários congelados. Descongele o cenário primeiro.'}), 400
        
        data = request.get_json()
        
        # Validações obrigatórias
        if not data.get('categoria_id'):
            return jsonify({'message': 'Categoria é obrigatória'}), 400
        if not data.get('data_competencia'):
            return jsonify({'message': 'Data de competência é obrigatória'}), 400
        if data.get('valor') is None:
            return jsonify({'message': 'Valor é obrigatório'}), 400
        if not data.get('tipo') or data.get('tipo') not in ['ENTRADA', 'SAIDA']:
            return jsonify({'message': 'Tipo deve ser ENTRADA ou SAIDA'}), 400
        
        # Validar categoria existe
        categoria = CategoriaFinanceira.query.get(data.get('categoria_id'))
        if not categoria:
            return jsonify({'message': 'Categoria não encontrada'}), 404
        
        # Validar data
        try:
            data_competencia = datetime.strptime(data.get('data_competencia'), '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'message': 'Formato de data inválido. Use YYYY-MM-DD'}), 400
        
        # Validar valor
        try:
            valor = float(data.get('valor'))
            if valor < 0:
                return jsonify({'message': 'Valor não pode ser negativo'}), 400
        except (ValueError, TypeError):
            return jsonify({'message': 'Valor inválido'}), 400
        
        # Criar lançamento
        lancamento = LancamentoFinanceiro(
            cenario_id=cenario_id,
            categoria_id=data.get('categoria_id'),
            data_competencia=data_competencia,
            valor=valor,
            tipo=data.get('tipo'),
            origem=data.get('origem', 'PROJETADO')
        )
        
        db.session.add(lancamento)
        db.session.commit()
        
        # Log
        log = LogSistema(
            usuario_id=current_user.id,
            acao='LANCAMENTO_CREATED',
            detalhes={
                'cenario_id': cenario_id,
                'lancamento_id': lancamento.id,
                'valor': float(valor),
                'tipo': data.get('tipo')
            }
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'message': 'Lançamento criado com sucesso',
            'lancamento': lancamento.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/cenarios/<int:cenario_id>/lancamentos/<int:lancamento_id>', methods=['PUT'])
@token_required
def atualizar_lancamento(current_user, cenario_id, lancamento_id):
    """Atualiza um lançamento financeiro"""
    try:
        cenario = Cenario.query.get_or_404(cenario_id)
        lancamento = LancamentoFinanceiro.query.filter_by(
            id=lancamento_id,
            cenario_id=cenario_id
        ).first_or_404()
        
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        # Validar que o cenário está ativo para edição
        if not cenario.is_active:
            return jsonify({'message': 'Não é possível editar lançamentos em cenários congelados. Descongele o cenário primeiro.'}), 400
        
        data = request.get_json()
        
        # Atualizar campos fornecidos
        if data.get('categoria_id'):
            categoria = CategoriaFinanceira.query.get(data.get('categoria_id'))
            if not categoria:
                return jsonify({'message': 'Categoria não encontrada'}), 404
            lancamento.categoria_id = data.get('categoria_id')
        
        if data.get('data_competencia'):
            try:
                lancamento.data_competencia = datetime.strptime(data.get('data_competencia'), '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'message': 'Formato de data inválido. Use YYYY-MM-DD'}), 400
        
        if data.get('valor') is not None:
            try:
                valor = float(data.get('valor'))
                if valor < 0:
                    return jsonify({'message': 'Valor não pode ser negativo'}), 400
                lancamento.valor = valor
            except (ValueError, TypeError):
                return jsonify({'message': 'Valor inválido'}), 400
        
        if data.get('tipo') and data.get('tipo') in ['ENTRADA', 'SAIDA']:
            lancamento.tipo = data.get('tipo')
        
        if data.get('origem') and data.get('origem') in ['PROJETADO', 'REALIZADO']:
            lancamento.origem = data.get('origem')
        
        db.session.commit()
        
        # Log
        log = LogSistema(
            usuario_id=current_user.id,
            acao='LANCAMENTO_UPDATED',
            detalhes={
                'cenario_id': cenario_id,
                'lancamento_id': lancamento_id,
                'alteracoes': data
            }
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'message': 'Lançamento atualizado com sucesso',
            'lancamento': lancamento.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/cenarios/<int:cenario_id>/lancamentos/<int:lancamento_id>', methods=['DELETE'])
@token_required
def deletar_lancamento(current_user, cenario_id, lancamento_id):
    """Deleta um lançamento financeiro"""
    try:
        cenario = Cenario.query.get_or_404(cenario_id)
        lancamento = LancamentoFinanceiro.query.filter_by(
            id=lancamento_id,
            cenario_id=cenario_id
        ).first_or_404()
        
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        # Validar que o cenário está ativo para edição
        if not cenario.is_active:
            return jsonify({'message': 'Não é possível deletar lançamentos em cenários congelados. Descongele o cenário primeiro.'}), 400
        
        # Log antes de deletar
        log = LogSistema(
            usuario_id=current_user.id,
            acao='LANCAMENTO_DELETED',
            detalhes={
                'cenario_id': cenario_id,
                'lancamento_id': lancamento_id,
                'valor': float(lancamento.valor) if lancamento.valor else 0,
                'tipo': lancamento.tipo
            }
        )
        db.session.add(log)
        
        db.session.delete(lancamento)
        db.session.commit()
        
        return jsonify({'message': 'Lançamento deletado com sucesso'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/cenarios/<int:cenario_id>/lancamentos', methods=['GET'])
@token_required
def listar_lancamentos_cenario(current_user, cenario_id):
    """Lista todos os lançamentos de um cenário"""
    try:
        cenario = Cenario.query.get_or_404(cenario_id)
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        # Buscar lançamentos com informações da categoria
        lancamentos = db.session.query(
            LancamentoFinanceiro,
            CategoriaFinanceira.nome.label('categoria_nome')
        ).join(
            CategoriaFinanceira, LancamentoFinanceiro.categoria_id == CategoriaFinanceira.id
        ).filter(
            LancamentoFinanceiro.cenario_id == cenario_id
        ).order_by(
            LancamentoFinanceiro.data_competencia.desc(),
            LancamentoFinanceiro.id.desc()
        ).all()
        
        lancamentos_data = []
        for lancamento, categoria_nome in lancamentos:
            lancamento_dict = lancamento.to_dict()
            lancamento_dict['categoria_nome'] = categoria_nome
            lancamentos_data.append(lancamento_dict)
        
        return jsonify({
            'lancamentos': lancamentos_data,
            'total': len(lancamentos_data)
        }), 200
        
    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/categorias', methods=['GET'])
@token_required
def listar_categorias(current_user):
    """Lista todas as categorias financeiras disponíveis"""
    try:
        categorias = CategoriaFinanceira.query.order_by(CategoriaFinanceira.nome).all()
        categorias_data = [cat.to_dict() for cat in categorias]
        return jsonify({'categorias': categorias_data}), 200
    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/cenarios/<int:cenario_id>/graficos', methods=['GET'])
@token_required
def obter_graficos_cenario(current_user, cenario_id):
    """Retorna dados agregados para gráficos de um cenário"""
    try:
        cenario = Cenario.query.get_or_404(cenario_id)
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        # Obter parâmetro de período (mensal, trimestral, anual)
        periodo = request.args.get('periodo', 'mensal')  # default: mensal
        
        # Buscar lançamentos do cenário
        lancamentos = LancamentoFinanceiro.query.filter_by(cenario_id=cenario_id).all()
        
        if not lancamentos:
            return jsonify({
                'fluxo_caixa': [],
                'distribuicao_categorias': [],
                'entradas_vs_saidas': [],
                'tendencias': {}
            }), 200
        
        # Agrupar dados por período
        dados_agrupados = {}
        
        for lancamento in lancamentos:
            data = lancamento.data_competencia
            
            if periodo == 'mensal':
                chave = f"{data.year}-{data.month:02d}"
                # Formatando nome do mês em português
                meses_pt = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
                periodo_label = f"{meses_pt[data.month - 1]}/{data.year}"
            elif periodo == 'trimestral':
                trimestre = (data.month - 1) // 3 + 1
                chave = f"{data.year}-T{trimestre}"
                periodo_label = f"{data.year} T{trimestre}"
            else:  # anual
                chave = str(data.year)
                periodo_label = str(data.year)
            
            if chave not in dados_agrupados:
                dados_agrupados[chave] = {
                    'periodo': periodo_label,
                    'entradas': 0,
                    'saidas': 0,
                    'saldo_liquido': 0
                }
            
            valor = float(lancamento.valor)
            if lancamento.tipo == 'ENTRADA':
                dados_agrupados[chave]['entradas'] += valor
            else:
                dados_agrupados[chave]['saidas'] += valor
            
            dados_agrupados[chave]['saldo_liquido'] = dados_agrupados[chave]['entradas'] - dados_agrupados[chave]['saidas']
        
        # Ordenar por período (usar a chave para ordenação correta)
        fluxo_caixa = [
            {
                'periodo': dados['periodo'],
                'entradas': dados['entradas'],
                'saidas': dados['saidas'],
                'saldo_liquido': dados['saldo_liquido']
            }
            for chave, dados in sorted(dados_agrupados.items())
        ]
        
        # Distribuição por categoria
        distribuicao_categorias = db.session.query(
            CategoriaFinanceira.nome,
            func.sum(LancamentoFinanceiro.valor).label('total'),
            LancamentoFinanceiro.tipo
        ).join(
            LancamentoFinanceiro, CategoriaFinanceira.id == LancamentoFinanceiro.categoria_id
        ).filter(
            LancamentoFinanceiro.cenario_id == cenario_id
        ).group_by(
            CategoriaFinanceira.nome,
            LancamentoFinanceiro.tipo
        ).all()
        
        categorias_dict = {}
        for categoria_nome, total, tipo in distribuicao_categorias:
            if categoria_nome not in categorias_dict:
                categorias_dict[categoria_nome] = {'entradas': 0, 'saidas': 0}
            
            valor_total = float(total) if total else 0
            if tipo == 'ENTRADA':
                categorias_dict[categoria_nome]['entradas'] += valor_total
            else:
                categorias_dict[categoria_nome]['saidas'] += valor_total
        
        distribuicao_categorias_data = [
            {
                'categoria': nome,
                'entradas': dados['entradas'],
                'saidas': dados['saidas'],
                'total': dados['entradas'] + dados['saidas']
            }
            for nome, dados in categorias_dict.items()
        ]
        
        # Ordenar por total (maior primeiro)
        distribuicao_categorias_data.sort(key=lambda x: x['total'], reverse=True)
        
        # Entradas vs Saídas (agregado por período)
        entradas_vs_saidas = [
            {
                'periodo': item['periodo'],
                'entradas': item['entradas'],
                'saidas': item['saidas']
            }
            for item in fluxo_caixa
        ]
        
        # Calcular tendências
        if len(fluxo_caixa) >= 2:
            ultimo_saldo = fluxo_caixa[-1]['saldo_liquido']
            penultimo_saldo = fluxo_caixa[-2]['saldo_liquido']
            
            variacao = ultimo_saldo - penultimo_saldo
            variacao_percentual = (variacao / penultimo_saldo * 100) if penultimo_saldo != 0 else 0
            
            # Calcular média de entradas e saídas
            media_entradas = sum(item['entradas'] for item in fluxo_caixa) / len(fluxo_caixa)
            media_saidas = sum(item['saidas'] for item in fluxo_caixa) / len(fluxo_caixa)
            
            # Projeção para próximo período (baseado na média)
            ultima_entrada = fluxo_caixa[-1]['entradas']
            ultima_saida = fluxo_caixa[-1]['saidas']
            
            # Tendência simples: média entre último valor e média histórica
            projecao_entrada = (ultima_entrada + media_entradas) / 2
            projecao_saida = (ultima_saida + media_saidas) / 2
            projecao_saldo = projecao_entrada - projecao_saida
            
            tendencias = {
                'variacao_saldo': variacao,
                'variacao_percentual': round(variacao_percentual, 2),
                'tendencia': 'crescente' if variacao > 0 else 'decrescente' if variacao < 0 else 'estavel',
                'media_entradas': round(media_entradas, 2),
                'media_saidas': round(media_saidas, 2),
                'projecao_entrada': round(projecao_entrada, 2),
                'projecao_saida': round(projecao_saida, 2),
                'projecao_saldo': round(projecao_saldo, 2)
            }
        else:
            tendencias = {
                'variacao_saldo': 0,
                'variacao_percentual': 0,
                'tendencia': 'insuficientes_dados',
                'media_entradas': 0,
                'media_saidas': 0,
                'projecao_entrada': 0,
                'projecao_saida': 0,
                'projecao_saldo': 0
            }
        
        return jsonify({
            'fluxo_caixa': fluxo_caixa,
            'distribuicao_categorias': distribuicao_categorias_data,
            'entradas_vs_saidas': entradas_vs_saidas,
            'tendencias': tendencias,
            'periodo': periodo
        }), 200
        
    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

# ============================================================================
# FUNÇÕES AUXILIARES PARA GERAÇÃO DE RELATÓRIOS
# ============================================================================

def _agrupar_por_periodo(lancamentos, periodo):
    """Agrupa lançamentos por período"""
    dados_agrupados = {}
    meses_pt = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    
    for lancamento in lancamentos:
        data = lancamento.data_competencia
        
        if periodo == 'mensal':
            chave = f"{data.year}-{data.month:02d}"
            periodo_label = f"{meses_pt[data.month - 1]}/{data.year}"
        elif periodo == 'trimestral':
            trimestre = (data.month - 1) // 3 + 1
            chave = f"{data.year}-T{trimestre}"
            periodo_label = f"{data.year} T{trimestre}"
        elif periodo == 'anual':
            chave = str(data.year)
            periodo_label = str(data.year)
        else:  # todos
            chave = 'todos'
            periodo_label = 'Todos os períodos'
        
        if chave not in dados_agrupados:
            dados_agrupados[chave] = {
                'periodo': periodo_label,
                'entradas': 0,
                'saidas': 0,
                'saldo_liquido': 0
            }
        
        valor = float(lancamento.valor)
        if lancamento.tipo == 'ENTRADA':
            dados_agrupados[chave]['entradas'] += valor
        else:
            dados_agrupados[chave]['saidas'] += valor
        
        dados_agrupados[chave]['saldo_liquido'] = dados_agrupados[chave]['entradas'] - dados_agrupados[chave]['saidas']
    
    return dados_agrupados


def _obter_top_categorias(lancamentos, top_n=5):
    """Retorna as top N categorias por valor total"""
    categorias_dict = {}
    
    for lancamento in lancamentos:
        categoria = CategoriaFinanceira.query.get(lancamento.categoria_id)
        categoria_nome = categoria.nome if categoria else 'N/A'
        
        if categoria_nome not in categorias_dict:
            categorias_dict[categoria_nome] = {
                'nome': categoria_nome,
                'entradas': 0,
                'saidas': 0,
                'total': 0
            }
        
        valor = float(lancamento.valor)
        if lancamento.tipo == 'ENTRADA':
            categorias_dict[categoria_nome]['entradas'] += valor
        else:
            categorias_dict[categoria_nome]['saidas'] += valor
        categorias_dict[categoria_nome]['total'] += valor if lancamento.tipo == 'ENTRADA' else -valor
    
    # Ordenar por valor total absoluto
    top_categorias = sorted(
        categorias_dict.values(),
        key=lambda x: abs(x['total']),
        reverse=True
    )[:top_n]
    
    return top_categorias


def _gerar_pdf_executive(cenario, projeto, lancamentos, periodo, upload_recente, styles, colors):
    """Gera PDF no template Executivo (resumido)"""
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.units import inch
    
    story = []
    
    # Calcular estatísticas
    total_entradas = sum(float(l.valor) for l in lancamentos if l.tipo == 'ENTRADA')
    total_saidas = sum(float(l.valor) for l in lancamentos if l.tipo == 'SAIDA')
    saldo_liquido = total_entradas - total_saidas
    
    # Título executivo
    title_style = ParagraphStyle(
        'ExecutiveTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    story.append(Paragraph("Relatório Executivo", title_style))
    story.append(Paragraph(f"<b>Cenário:</b> {cenario.nome}", styles['Heading2']))
    story.append(Spacer(1, 0.3*inch))
    
    # Resumo financeiro em destaque
    summary_data = [
        ['Indicador', 'Valor'],
        ['Total de Entradas', f"R$ {total_entradas:,.2f}"],
        ['Total de Saídas', f"R$ {total_saidas:,.2f}"],
        ['Saldo Líquido', f"R$ {saldo_liquido:,.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[4*inch, 4*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f9ff'))
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.4*inch))
    
    # Informações essenciais
    info_data = [
        ['Projeto:', projeto.nome_cliente],
        ['Status:', 'Ativo' if cenario.is_active else 'Congelado'],
    ]
    if upload_recente:
        info_data.append(['Arquivo Origem:', upload_recente.nome_original])
    
    info_table = Table(info_data, colWidths=[2.5*inch, 6*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(info_table)
    
    return story


def _gerar_pdf_detailed(cenario, projeto, lancamentos, periodo, upload_recente, styles, colors):
    """Gera PDF no template Detalhado (completo)"""
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.units import inch
    
    story = []
    
    # Calcular estatísticas
    total_entradas = sum(float(l.valor) for l in lancamentos if l.tipo == 'ENTRADA')
    total_saidas = sum(float(l.valor) for l in lancamentos if l.tipo == 'SAIDA')
    saldo_liquido = total_entradas - total_saidas
    total_lancamentos = len(lancamentos)
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#374151'),
        spaceAfter=8
    )
    
    story.append(Paragraph(f"Relatório Detalhado do Cenário: {cenario.nome}", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Informações do Projeto
    story.append(Paragraph("Informações do Projeto", heading_style))
    info_data = [
        ['Projeto:', projeto.nome_cliente],
        ['Data Base:', projeto.data_base_estudo.strftime('%d/%m/%Y') if projeto.data_base_estudo else 'N/A'],
        ['Saldo Inicial:', f"R$ {float(projeto.saldo_inicial_caixa):,.2f}"],
        ['Status:', 'Ativo' if cenario.is_active else 'Congelado'],
        ['Arquivo Origem:', upload_recente.nome_original if upload_recente else 'N/A']
    ]
    info_table = Table(info_data, colWidths=[2.5*inch, 6*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Estatísticas Financeiras
    story.append(Paragraph("Estatísticas Financeiras", heading_style))
    stats_data = [
        ['Total de Entradas', f"R$ {total_entradas:,.2f}"],
        ['Total de Saídas', f"R$ {total_saidas:,.2f}"],
        ['Saldo Líquido', f"R$ {saldo_liquido:,.2f}"],
        ['Total de Lançamentos', str(total_lancamentos)]
    ]
    stats_table = Table(stats_data, colWidths=[4*inch, 4*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Top 5 Categorias
    top_categorias = _obter_top_categorias(lancamentos, 5)
    if top_categorias:
        story.append(Paragraph("Top 5 Categorias", heading_style))
        cat_data = [['Categoria', 'Entradas', 'Saídas', 'Total']]
        for cat in top_categorias:
            cat_data.append([
                cat['nome'],
                f"R$ {cat['entradas']:,.2f}",
                f"R$ {cat['saidas']:,.2f}",
                f"R$ {cat['total']:,.2f}"
            ])
        cat_table = Table(cat_data, colWidths=[3*inch, 2*inch, 2*inch, 2*inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        story.append(cat_table)
        story.append(Spacer(1, 0.3*inch))
    
    # Fluxo por período
    if periodo != 'todos':
        dados_agrupados = _agrupar_por_periodo(lancamentos, periodo)
        if dados_agrupados:
            story.append(Paragraph(f"Fluxo de Caixa por Período ({periodo})", heading_style))
            periodo_data = [['Período', 'Entradas', 'Saídas', 'Saldo']]
            for chave in sorted(dados_agrupados.keys()):
                dados = dados_agrupados[chave]
                periodo_data.append([
                    dados['periodo'],
                    f"R$ {dados['entradas']:,.2f}",
                    f"R$ {dados['saidas']:,.2f}",
                    f"R$ {dados['saldo_liquido']:,.2f}"
                ])
            
            periodo_table = Table(periodo_data, colWidths=[2*inch, 2*inch, 2*inch, 2*inch])
            periodo_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
            ]))
            story.append(periodo_table)
    
    return story


def _gerar_pdf_comparison(cenarios_data, periodo, styles, colors):
    """Gera PDF no template Comparativo (múltiplos cenários)"""
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.units import inch
    
    story = []
    
    # Título comparativo
    title_style = ParagraphStyle(
        'ComparisonTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'ComparisonHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#374151'),
        spaceAfter=8
    )
    
    story.append(Paragraph("Relatório Comparativo de Cenários", title_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Tabela comparativa de estatísticas
    story.append(Paragraph("Comparativo de Estatísticas", heading_style))
    
    # Cabeçalho
    comp_headers = ['Cenário', 'Projeto', 'Entradas', 'Saídas', 'Saldo Líquido', 'Lançamentos']
    comp_data = [comp_headers]
    
    for item in cenarios_data:
        cenario = item['cenario']
        projeto = item['projeto']
        lancamentos = item['lancamentos']
        
        total_entradas = sum(float(l.valor) for l in lancamentos if l.tipo == 'ENTRADA')
        total_saidas = sum(float(l.valor) for l in lancamentos if l.tipo == 'SAIDA')
        saldo_liquido = total_entradas - total_saidas
        
        comp_data.append([
            cenario.nome,
            projeto.nome_cliente,
            f"R$ {total_entradas:,.2f}",
            f"R$ {total_saidas:,.2f}",
            f"R$ {saldo_liquido:,.2f}",
            str(len(lancamentos))
        ])
    
    comp_table = Table(comp_data, colWidths=[2*inch, 2.5*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1.2*inch])
    comp_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
    ]))
    story.append(comp_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Análise individual de cada cenário
    for idx, item in enumerate(cenarios_data):
        if idx > 0:
            story.append(PageBreak())
        
        cenario = item['cenario']
        projeto = item['projeto']
        lancamentos = item['lancamentos']
        
        story.append(Paragraph(f"Cenário: {cenario.nome}", heading_style))
        story.append(Paragraph(f"Projeto: {projeto.nome_cliente}", styles['Normal']))
        story.append(Spacer(1, 0.2*inch))
        
        # Top 3 categorias deste cenário
        top_categorias = _obter_top_categorias(lancamentos, 3)
        if top_categorias:
            story.append(Paragraph("Top 3 Categorias", styles['Heading3']))
            cat_data = [['Categoria', 'Total']]
            for cat in top_categorias:
                cat_data.append([
                    cat['nome'],
                    f"R$ {cat['total']:,.2f}"
                ])
            cat_table = Table(cat_data, colWidths=[4*inch, 3*inch])
            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6b7280')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            story.append(cat_table)
    
    return story


@projetos_bp.route('/cenarios/<int:cenario_id>/relatorio/pdf', methods=['GET'])
@token_required
def gerar_relatorio_pdf(current_user, cenario_id):
    """Gera relatório PDF de um cenário"""
    try:
        from reportlab.lib.pagesizes import letter, A4, landscape
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        
        cenario = Cenario.query.get_or_404(cenario_id)
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        # Obter parâmetros
        periodo = request.args.get('periodo', 'todos')  # todos, mensal, trimestral, anual
        template = request.args.get('template', 'detailed')  # executive, detailed, comparison
        
        # Buscar lançamentos do cenário
        lancamentos_query = LancamentoFinanceiro.query.filter_by(cenario_id=cenario_id)
        
        # Aplicar filtro de período se não for 'todos'
        if periodo != 'todos':
            hoje = datetime.now().date()
            if periodo == 'mensal':
                # Último mês
                primeiro_dia_mes = hoje.replace(day=1)
                lancamentos_query = lancamentos_query.filter(
                    LancamentoFinanceiro.data_competencia >= primeiro_dia_mes
                )
            elif periodo == 'trimestral':
                # Último trimestre
                trimestre_atual = (hoje.month - 1) // 3
                primeiro_mes_trimestre = trimestre_atual * 3 + 1
                primeiro_dia_trimestre = hoje.replace(month=primeiro_mes_trimestre, day=1)
                lancamentos_query = lancamentos_query.filter(
                    LancamentoFinanceiro.data_competencia >= primeiro_dia_trimestre
                )
            elif periodo == 'anual':
                # Último ano
                primeiro_dia_ano = hoje.replace(month=1, day=1)
                lancamentos_query = lancamentos_query.filter(
                    LancamentoFinanceiro.data_competencia >= primeiro_dia_ano
                )
        
        lancamentos = lancamentos_query.all()
        
        # Verificar se há lançamentos
        if not lancamentos:
            return jsonify({'message': 'Não há lançamentos para gerar o relatório. Por favor, adicione lançamentos ao cenário primeiro.'}), 400
        
        # Calcular estatísticas
        total_entradas = sum(float(l.valor) for l in lancamentos if l.tipo == 'ENTRADA')
        total_saidas = sum(float(l.valor) for l in lancamentos if l.tipo == 'SAIDA')
        saldo_liquido = total_entradas - total_saidas
        total_lancamentos = len(lancamentos)
        
        # Buscar arquivo relacionado
        upload_recente = ArquivoUpload.query.filter_by(projeto_id=projeto.id)\
            .order_by(ArquivoUpload.uploaded_at.desc()).first()
        
        # Criar PDF em memória (paisagem)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Validar template
        if template not in ['executive', 'detailed', 'comparison']:
            template = 'detailed'
        
        # Gerar conteúdo baseado no template
        if template == 'executive':
            story = _gerar_pdf_executive(cenario, projeto, lancamentos, periodo, upload_recente, styles, colors)
        elif template == 'detailed':
            story = _gerar_pdf_detailed(cenario, projeto, lancamentos, periodo, upload_recente, styles, colors)
        else:
            # Comparison requer múltiplos cenários - usar detailed como fallback
            story = _gerar_pdf_detailed(cenario, projeto, lancamentos, periodo, upload_recente, styles, colors)
        
        # Adicionar data de geração e rodapé
        from reportlab.platypus import Spacer
        story.append(Spacer(1, 0.2*inch))
        story.append(Paragraph(
            f"<i>Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}</i>",
            styles['Normal']
        ))
        
        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        
        # Contar páginas do PDF usando PyPDF2
        num_pages = 1  # Default
        try:
            import PyPDF2
            pdf_reader = PyPDF2.PdfReader(buffer)
            num_pages = len(pdf_reader.pages)
            buffer.seek(0)  # Resetar buffer após leitura
        except ImportError:
            # Se PyPDF2 não estiver instalado, estimar baseado no tamanho do conteúdo
            # Estimativa: ~50 linhas por página
            total_elements = len(story)
            num_pages = max(1, (total_elements // 50) + 1)
        except Exception:
            # Em caso de erro, usar estimativa
            total_elements = len(story)
            num_pages = max(1, (total_elements // 50) + 1)
        
        # Preparar nome do arquivo
        nome_arquivo = f"relatorio_{template}_{cenario.nome.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        response = send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=nome_arquivo
        )
        
        # Adicionar metadados nos headers
        response.headers['X-Report-Pages'] = str(num_pages)
        response.headers['X-Report-Template'] = template
        response.headers['X-Report-Period'] = periodo
        
        return response
        
    except Exception as e:
        return jsonify({'message': f'Erro ao gerar PDF: {str(e)}'}), 500


@projetos_bp.route('/cenarios/relatorio-comparativo/pdf', methods=['POST'])
@token_required
def gerar_relatorio_comparativo_pdf(current_user):
    """Gera relatório PDF comparativo de múltiplos cenários"""
    try:
        from reportlab.lib.pagesizes import letter, A4, landscape
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        
        data = request.get_json()
        
        if not data or not data.get('cenario_ids'):
            return jsonify({'message': 'Lista de IDs de cenários é obrigatória'}), 400
        
        cenario_ids = data.get('cenario_ids')
        periodo = data.get('periodo', 'todos')
        
        if not isinstance(cenario_ids, list) or len(cenario_ids) < 2:
            return jsonify({'message': 'É necessário pelo menos 2 cenários para comparar'}), 400
        
        # Buscar cenários e verificar permissões
        cenarios_data = []
        for cenario_id in cenario_ids:
            cenario = Cenario.query.get(cenario_id)
            if not cenario:
                return jsonify({'message': f'Cenário {cenario_id} não encontrado'}), 404
            
            projeto = Projeto.query.get(cenario.projeto_id)
            if not projeto:
                return jsonify({'message': f'Projeto do cenário {cenario_id} não encontrado'}), 404
            
            # Verificar permissão
            if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
                return jsonify({'message': f'Acesso negado ao cenário {cenario_id}'}), 403
            
            # Aplicar filtro de período
            lancamentos_query = LancamentoFinanceiro.query.filter_by(cenario_id=cenario.id)
            
            if periodo != 'todos':
                hoje = datetime.now().date()
                if periodo == 'mensal':
                    primeiro_dia_mes = hoje.replace(day=1)
                    lancamentos_query = lancamentos_query.filter(
                        LancamentoFinanceiro.data_competencia >= primeiro_dia_mes
                    )
                elif periodo == 'trimestral':
                    trimestre_atual = (hoje.month - 1) // 3
                    primeiro_mes_trimestre = trimestre_atual * 3 + 1
                    primeiro_dia_trimestre = hoje.replace(month=primeiro_mes_trimestre, day=1)
                    lancamentos_query = lancamentos_query.filter(
                        LancamentoFinanceiro.data_competencia >= primeiro_dia_trimestre
                    )
                elif periodo == 'anual':
                    primeiro_dia_ano = hoje.replace(month=1, day=1)
                    lancamentos_query = lancamentos_query.filter(
                        LancamentoFinanceiro.data_competencia >= primeiro_dia_ano
                    )
            
            lancamentos = lancamentos_query.all()
            
            cenarios_data.append({
                'cenario': cenario,
                'projeto': projeto,
                'lancamentos': lancamentos
            })
        
        # Criar PDF em memória (paisagem)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        
        # Gerar conteúdo comparativo
        story = _gerar_pdf_comparison(cenarios_data, periodo, styles, colors)
        
        # Adicionar data de geração
        from reportlab.platypus import Spacer
        story.append(Spacer(1, 0.2*inch))
        story.append(Paragraph(
            f"<i>Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}</i>",
            styles['Normal']
        ))
        
        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        
        # Contar páginas do PDF
        num_pages = 1
        try:
            import PyPDF2
            pdf_reader = PyPDF2.PdfReader(buffer)
            num_pages = len(pdf_reader.pages)
            buffer.seek(0)
        except ImportError:
            # Estimativa baseada no número de cenários e conteúdo
            num_pages = max(1, len(cenarios_data) + 1)
        except Exception:
            num_pages = max(1, len(cenarios_data) + 1)
        
        # Preparar nome do arquivo
        nomes_cenarios = '_'.join([c['cenario'].nome.replace(' ', '_') for c in cenarios_data[:3]])
        nome_arquivo = f"relatorio_comparativo_{nomes_cenarios}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        response = send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=nome_arquivo
        )
        
        # Adicionar metadados nos headers
        response.headers['X-Report-Pages'] = str(num_pages)
        response.headers['X-Report-Template'] = 'comparison'
        response.headers['X-Report-Period'] = periodo
        
        return response
        
    except Exception as e:
        return jsonify({'message': f'Erro ao gerar PDF comparativo: {str(e)}'}), 500


@projetos_bp.route('/cenarios/relatorio-comparativo/excel', methods=['POST'])
@token_required
def gerar_relatorio_comparativo_excel(current_user):
    """Gera relatório Excel comparativo de múltiplos cenários"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        data = request.get_json()
        
        if not data or not data.get('cenario_ids'):
            return jsonify({'message': 'Lista de IDs de cenários é obrigatória'}), 400
        
        cenario_ids = data.get('cenario_ids')
        periodo = data.get('periodo', 'todos')
        
        if not isinstance(cenario_ids, list) or len(cenario_ids) < 2:
            return jsonify({'message': 'É necessário pelo menos 2 cenários para comparar'}), 400
        
        # Buscar cenários e verificar permissões
        cenarios_data = []
        for cenario_id in cenario_ids:
            cenario = Cenario.query.get(cenario_id)
            if not cenario:
                return jsonify({'message': f'Cenário {cenario_id} não encontrado'}), 404
            
            projeto = Projeto.query.get(cenario.projeto_id)
            if not projeto:
                return jsonify({'message': f'Projeto do cenário {cenario_id} não encontrado'}), 404
            
            # Verificar permissão
            if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
                return jsonify({'message': f'Acesso negado ao cenário {cenario_id}'}), 403
            
            # Aplicar filtro de período
            lancamentos_query = LancamentoFinanceiro.query.filter_by(cenario_id=cenario.id)
            
            if periodo != 'todos':
                hoje = datetime.now().date()
                if periodo == 'mensal':
                    primeiro_dia_mes = hoje.replace(day=1)
                    lancamentos_query = lancamentos_query.filter(
                        LancamentoFinanceiro.data_competencia >= primeiro_dia_mes
                    )
                elif periodo == 'trimestral':
                    trimestre_atual = (hoje.month - 1) // 3
                    primeiro_mes_trimestre = trimestre_atual * 3 + 1
                    primeiro_dia_trimestre = hoje.replace(month=primeiro_mes_trimestre, day=1)
                    lancamentos_query = lancamentos_query.filter(
                        LancamentoFinanceiro.data_competencia >= primeiro_dia_trimestre
                    )
                elif periodo == 'anual':
                    primeiro_dia_ano = hoje.replace(month=1, day=1)
                    lancamentos_query = lancamentos_query.filter(
                        LancamentoFinanceiro.data_competencia >= primeiro_dia_ano
                    )
            
            lancamentos = lancamentos_query.all()
            
            cenarios_data.append({
                'cenario': cenario,
                'projeto': projeto,
                'lancamentos': lancamentos
            })
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativo"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Título
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Relatório Comparativo de Cenários"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        # Tabela comparativa
        ws[f'A{row}'] = "Cenário"
        ws[f'B{row}'] = "Projeto"
        ws[f'C{row}'] = "Entradas"
        ws[f'D{row}'] = "Saídas"
        ws[f'E{row}'] = "Saldo Líquido"
        ws[f'F{row}'] = "Lançamentos"
        
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Dados dos cenários
        for item in cenarios_data:
            cenario = item['cenario']
            projeto = item['projeto']
            lancamentos = item['lancamentos']
            
            total_entradas = sum(float(l.valor) for l in lancamentos if l.tipo == 'ENTRADA')
            total_saidas = sum(float(l.valor) for l in lancamentos if l.tipo == 'SAIDA')
            saldo_liquido = total_entradas - total_saidas
            
            ws[f'A{row}'] = cenario.nome
            ws[f'B{row}'] = projeto.nome_cliente
            ws[f'C{row}'] = f"R$ {total_entradas:,.2f}"
            ws[f'D{row}'] = f"R$ {total_saidas:,.2f}"
            ws[f'E{row}'] = f"R$ {saldo_liquido:,.2f}"
            ws[f'F{row}'] = len(lancamentos)
            
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col >= 3:
                    cell.alignment = Alignment(horizontal='right')
            
            row += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        
        # Contar número de planilhas
        num_sheets = len(wb.worksheets)
        
        # Criar arquivo em memória
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Nome do arquivo
        nomes_cenarios = '_'.join([c['cenario'].nome.replace(' ', '_') for c in cenarios_data[:3]])
        nome_arquivo = f"relatorio_comparativo_{nomes_cenarios}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )
        
        # Adicionar metadados nos headers
        response.headers['X-Report-Sheets'] = str(num_sheets)
        response.headers['X-Report-Template'] = 'comparison'
        response.headers['X-Report-Period'] = periodo
        
        return response
        
    except Exception as e:
        return jsonify({'message': f'Erro ao gerar Excel comparativo: {str(e)}'}), 500


@projetos_bp.route('/cenarios/<int:cenario_id>/relatorio/excel', methods=['GET'])
@token_required
def gerar_relatorio_excel(current_user, cenario_id):
    """Gera relatório Excel de um cenário"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        cenario = Cenario.query.get_or_404(cenario_id)
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        # Obter parâmetros
        periodo = request.args.get('periodo', 'todos')
        template = request.args.get('template', 'detailed')  # executive, detailed, comparison
        
        # Buscar lançamentos do cenário com filtro de período
        lancamentos_query = LancamentoFinanceiro.query.filter_by(cenario_id=cenario_id)
        
        # Aplicar filtro de período se não for 'todos'
        if periodo != 'todos':
            hoje = datetime.now().date()
            if periodo == 'mensal':
                primeiro_dia_mes = hoje.replace(day=1)
                lancamentos_query = lancamentos_query.filter(
                    LancamentoFinanceiro.data_competencia >= primeiro_dia_mes
                )
            elif periodo == 'trimestral':
                trimestre_atual = (hoje.month - 1) // 3
                primeiro_mes_trimestre = trimestre_atual * 3 + 1
                primeiro_dia_trimestre = hoje.replace(month=primeiro_mes_trimestre, day=1)
                lancamentos_query = lancamentos_query.filter(
                    LancamentoFinanceiro.data_competencia >= primeiro_dia_trimestre
                )
            elif periodo == 'anual':
                primeiro_dia_ano = hoje.replace(month=1, day=1)
                lancamentos_query = lancamentos_query.filter(
                    LancamentoFinanceiro.data_competencia >= primeiro_dia_ano
                )
        
        lancamentos = lancamentos_query.order_by(LancamentoFinanceiro.data_competencia).all()
        
        # Verificar se há lançamentos
        if not lancamentos:
            return jsonify({'message': 'Não há lançamentos para gerar o relatório. Por favor, adicione lançamentos ao cenário primeiro.'}), 400
        
        # Calcular estatísticas
        total_entradas = sum(float(l.valor) for l in lancamentos if l.tipo == 'ENTRADA')
        total_saidas = sum(float(l.valor) for l in lancamentos if l.tipo == 'SAIDA')
        saldo_liquido = total_entradas - total_saidas
        
        # Buscar arquivo relacionado
        upload_recente = ArquivoUpload.query.filter_by(projeto_id=projeto.id)\
            .order_by(ArquivoUpload.uploaded_at.desc()).first()
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Título
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = f"Relatório do Cenário: {cenario.nome}"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        # Informações do Projeto
        ws[f'A{row}'] = "Informações do Projeto"
        ws[f'A{row}'].font = title_font
        row += 1
        
        info_labels = ['Projeto:', 'Data Base:', 'Saldo Inicial:', 'Status:', 'Arquivo Origem:']
        info_values = [
            projeto.nome_cliente,
            projeto.data_base_estudo.strftime('%d/%m/%Y') if projeto.data_base_estudo else 'N/A',
            f"R$ {float(projeto.saldo_inicial_caixa):,.2f}",
            'Ativo' if cenario.is_active else 'Congelado',
            upload_recente.nome_original if upload_recente else 'N/A'
        ]
        
        for i, (label, value) in enumerate(zip(info_labels, info_values)):
            ws[f'A{row + i}'] = label
            ws[f'A{row + i}'].font = Font(bold=True)
            ws[f'B{row + i}'] = value
        row += len(info_labels) + 2
        
        # Estatísticas
        ws[f'A{row}'] = "Estatísticas Financeiras"
        ws[f'A{row}'].font = title_font
        row += 1
        
        stats_headers = ['Métrica', 'Valor']
        stats_data = [
            ['Total de Entradas', f"R$ {total_entradas:,.2f}"],
            ['Total de Saídas', f"R$ {total_saidas:,.2f}"],
            ['Saldo Líquido', f"R$ {saldo_liquido:,.2f}"],
            ['Total de Lançamentos', len(lancamentos)]
        ]
        
        # Cabeçalho
        for col, header in enumerate(stats_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Dados
        for stats_row in stats_data:
            for col, value in enumerate(stats_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 1:
                    cell.font = Font(bold=True)
            row += 1
        row += 1
        
        # Lançamentos
        ws[f'A{row}'] = "Lançamentos Detalhados"
        ws[f'A{row}'].font = title_font
        row += 1
        
        lanc_headers = ['Data', 'Categoria', 'Tipo', 'Valor', 'Origem']
        for col, header in enumerate(lanc_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Dados dos lançamentos
        for lancamento in lancamentos:
            categoria = CategoriaFinanceira.query.get(lancamento.categoria_id)
            categoria_nome = categoria.nome if categoria else 'N/A'
            
            lanc_data = [
                lancamento.data_competencia.strftime('%d/%m/%Y'),
                categoria_nome,
                lancamento.tipo,
                f"R$ {float(lancamento.valor):,.2f}",
                lancamento.origem
            ]
            
            for col, value in enumerate(lanc_data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 4:  # Valor
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        
        # Contar número de planilhas
        num_sheets = len(wb.worksheets)
        
        # Criar arquivo em memória
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Nome do arquivo
        nome_arquivo = f"relatorio_{cenario.nome.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )
        
        # Adicionar metadados nos headers
        response.headers['X-Report-Sheets'] = str(num_sheets)
        response.headers['X-Report-Template'] = template
        response.headers['X-Report-Period'] = periodo
        
        return response
        
    except Exception as e:
        return jsonify({'message': f'Erro ao gerar Excel: {str(e)}'}), 500

@projetos_bp.route('/cenarios/<int:cenario_id>/snapshot', methods=['POST'])
@token_required
def criar_snapshot_cenario(current_user, cenario_id):
    """Cria um snapshot (versão) do cenário atual"""
    try:
        cenario = Cenario.query.get_or_404(cenario_id)
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        data = request.get_json()
        descricao = data.get('descricao', '') if data else ''
        
        # Buscar todos os lançamentos do cenário
        lancamentos = LancamentoFinanceiro.query.filter_by(cenario_id=cenario_id).all()
        
        # Criar snapshot com dados serializados
        snapshot_data = {
            'cenario': {
                'nome': cenario.nome,
                'descricao': cenario.descricao,
                'is_active': cenario.is_active
            },
            'lancamentos': [
                {
                    'categoria_id': l.categoria_id,
                    'data_competencia': l.data_competencia.isoformat() if l.data_competencia else None,
                    'valor': float(l.valor) if l.valor else 0,
                    'tipo': l.tipo,
                    'origem': l.origem
                }
                for l in lancamentos
            ],
            'total_lancamentos': len(lancamentos),
            'total_entradas': sum(float(l.valor) for l in lancamentos if l.tipo == 'ENTRADA'),
            'total_saidas': sum(float(l.valor) for l in lancamentos if l.tipo == 'SAIDA'),
            'saldo_liquido': sum(float(l.valor) for l in lancamentos if l.tipo == 'ENTRADA') - 
                            sum(float(l.valor) for l in lancamentos if l.tipo == 'SAIDA')
        }
        
        # Criar registro de histórico
        historico = HistoricoCenario(
            cenario_id=cenario_id,
            usuario_id=current_user.id,
            descricao=descricao,
            snapshot_data=snapshot_data
        )
        
        db.session.add(historico)
        
        # Log
        log = LogSistema(
            usuario_id=current_user.id,
            acao='SNAPSHOT_CREATED',
            detalhes={
                'cenario_id': cenario_id,
                'historico_id': historico.id,
                'descricao': descricao
            }
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'message': 'Snapshot criado com sucesso',
            'historico': historico.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/cenarios/<int:cenario_id>/historico', methods=['GET'])
@token_required
def listar_historico_cenario(current_user, cenario_id):
    """Lista o histórico de versões de um cenário"""
    try:
        cenario = Cenario.query.get_or_404(cenario_id)
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        # Buscar histórico
        historicos = HistoricoCenario.query.filter_by(cenario_id=cenario_id)\
            .order_by(HistoricoCenario.created_at.desc()).all()
        
        historicos_data = []
        for historico in historicos:
            h_dict = historico.to_dict()
            # Adicionar informações do usuário
            usuario = db.session.get(User, historico.usuario_id)
            if usuario:
                h_dict['usuario_nome'] = usuario.nome
                h_dict['usuario_email'] = usuario.email
            historicos_data.append(h_dict)
        
        return jsonify({
            'historico': historicos_data,
            'total': len(historicos_data)
        }), 200
        
    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/cenarios/<int:cenario_id>/restaurar/<int:historico_id>', methods=['POST'])
@token_required
def restaurar_versao_cenario(current_user, cenario_id, historico_id):
    """Restaura um cenário para uma versão anterior"""
    try:
        cenario = Cenario.query.get_or_404(cenario_id)
        projeto = Projeto.query.get(cenario.projeto_id)
        
        if not projeto:
            return jsonify({'message': 'Projeto não encontrado'}), 404
        
        # Verificar permissão
        if current_user.role != 'admin' and projeto.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        # Validar que o cenário está ativo para edição
        if not cenario.is_active:
            return jsonify({'message': 'Não é possível restaurar versões em cenários congelados. Descongele o cenário primeiro.'}), 400
        
        # Buscar histórico
        historico = HistoricoCenario.query.filter_by(
            id=historico_id,
            cenario_id=cenario_id
        ).first_or_404()
        
        snapshot_data = historico.snapshot_data
        
        # Criar snapshot do estado atual antes de restaurar
        lancamentos_atuais = LancamentoFinanceiro.query.filter_by(cenario_id=cenario_id).all()
        snapshot_atual = {
            'cenario': {
                'nome': cenario.nome,
                'descricao': cenario.descricao,
                'is_active': cenario.is_active
            },
            'lancamentos': [
                {
                    'categoria_id': l.categoria_id,
                    'data_competencia': l.data_competencia.isoformat() if l.data_competencia else None,
                    'valor': float(l.valor) if l.valor else 0,
                    'tipo': l.tipo,
                    'origem': l.origem
                }
                for l in lancamentos_atuais
            ]
        }
        
        # Criar backup automático antes de restaurar
        backup = HistoricoCenario(
            cenario_id=cenario_id,
            usuario_id=current_user.id,
            descricao=f'Backup automático antes de restaurar versão de {historico.created_at.strftime("%d/%m/%Y %H:%M")}',
            snapshot_data=snapshot_atual
        )
        db.session.add(backup)
        
        # Deletar lançamentos atuais
        for lancamento in lancamentos_atuais:
            db.session.delete(lancamento)
        
        # Restaurar lançamentos do snapshot
        lancamentos_restaurados = 0
        for lanc_data in snapshot_data.get('lancamentos', []):
            try:
                # Verificar se a categoria ainda existe
                categoria = CategoriaFinanceira.query.get(lanc_data['categoria_id'])
                if not categoria:
                    continue  # Pular se categoria não existir mais
                
                from datetime import datetime as dt
                data_competencia = dt.fromisoformat(lanc_data['data_competencia']).date()
                
                lancamento = LancamentoFinanceiro(
                    cenario_id=cenario_id,
                    categoria_id=lanc_data['categoria_id'],
                    data_competencia=data_competencia,
                    valor=lanc_data['valor'],
                    tipo=lanc_data['tipo'],
                    origem=lanc_data['origem']
                )
                db.session.add(lancamento)
                lancamentos_restaurados += 1
            except Exception as e:
                print(f"Erro ao restaurar lançamento: {e}")
                continue
        
        # Restaurar dados do cenário (se houver mudanças)
        if snapshot_data.get('cenario'):
            cenario.nome = snapshot_data['cenario'].get('nome', cenario.nome)
            cenario.descricao = snapshot_data['cenario'].get('descricao', cenario.descricao)
        
        # Log
        log = LogSistema(
            usuario_id=current_user.id,
            acao='SCENARIO_RESTORED',
            detalhes={
                'cenario_id': cenario_id,
                'historico_id': historico_id,
                'backup_id': backup.id,
                'lancamentos_restaurados': lancamentos_restaurados
            }
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'message': 'Versão restaurada com sucesso',
            'lancamentos_restaurados': lancamentos_restaurados,
            'backup_id': backup.id
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

# ============================================================================
# ENDPOINTS PARA GERENCIAMENTO DE RELATÓRIOS
# ============================================================================

@projetos_bp.route('/relatorios', methods=['GET'])
@token_required
def listar_relatorios(current_user):
    """Lista todos os relatórios do usuário atual"""
    try:
        if current_user.role == 'admin':
            relatorios = Relatorio.query.order_by(Relatorio.created_at.desc()).all()
        else:
            relatorios = Relatorio.query.filter_by(usuario_id=current_user.id)\
                .order_by(Relatorio.created_at.desc()).all()
        
        return jsonify({
            'relatorios': [r.to_dict() for r in relatorios],
            'total': len(relatorios)
        }), 200
        
    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/relatorios', methods=['POST'])
@token_required
def criar_relatorio(current_user):
    """Cria um novo registro de relatório"""
    try:
        data = request.get_json()
        
        if not data or not data.get('title'):
            return jsonify({'message': 'Título do relatório é obrigatório'}), 400
        
        # Validar campos obrigatórios
        if not data.get('type') or data.get('type') not in ['pdf', 'excel']:
            return jsonify({'message': 'Tipo deve ser pdf ou excel'}), 400
        
        if not data.get('template') or data.get('template') not in ['executive', 'detailed', 'comparison']:
            return jsonify({'message': 'Template inválido'}), 400
        
        # Criar relatório
        relatorio = Relatorio(
            usuario_id=current_user.id,
            title=data.get('title'),
            type=data.get('type'),
            template=data.get('template'),
            scenario=data.get('scenario'),
            scenario_id=data.get('scenarioId'),
            scenario_ids=data.get('scenarioIds'),
            size=data.get('size'),
            pages=int(data.get('pages')) if data.get('pages') else None,
            sheets=int(data.get('sheets')) if data.get('sheets') else None,
            downloads=data.get('downloads', 1),
            status=data.get('status', 'completed'),
            periodo=data.get('periodo', 'todos'),
            descricao=data.get('descricao')
        )
        
        db.session.add(relatorio)
        db.session.commit()
        
        # Log
        log = LogSistema(
            usuario_id=current_user.id,
            acao='REPORT_CREATED',
            detalhes={
                'relatorio_id': relatorio.id,
                'title': relatorio.title,
                'type': relatorio.type
            }
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'message': 'Relatório criado com sucesso',
            'relatorio': relatorio.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/relatorios/<int:relatorio_id>', methods=['GET'])
@token_required
def obter_relatorio(current_user, relatorio_id):
    """Obtém um relatório específico"""
    try:
        relatorio = Relatorio.query.get_or_404(relatorio_id)
        
        # Verificar permissão
        if current_user.role != 'admin' and relatorio.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        return jsonify({'relatorio': relatorio.to_dict()}), 200
        
    except Exception as e:
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/relatorios/<int:relatorio_id>', methods=['PUT'])
@token_required
def atualizar_relatorio(current_user, relatorio_id):
    """Atualiza um relatório (ex: incrementar downloads)"""
    try:
        relatorio = Relatorio.query.get_or_404(relatorio_id)
        
        # Verificar permissão
        if current_user.role != 'admin' and relatorio.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        data = request.get_json()
        
        # Atualizar campos permitidos
        if data.get('downloads') is not None:
            relatorio.downloads = data.get('downloads')
        if data.get('title'):
            relatorio.title = data.get('title')
        if data.get('descricao') is not None:
            relatorio.descricao = data.get('descricao')
        
        relatorio.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'message': 'Relatório atualizado com sucesso',
            'relatorio': relatorio.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500

@projetos_bp.route('/relatorios/<int:relatorio_id>', methods=['DELETE'])
@token_required
def deletar_relatorio(current_user, relatorio_id):
    """Deleta um relatório"""
    try:
        relatorio = Relatorio.query.get_or_404(relatorio_id)
        
        # Verificar permissão
        if current_user.role != 'admin' and relatorio.usuario_id != current_user.id:
            return jsonify({'message': 'Acesso negado'}), 403
        
        title = relatorio.title
        
        # Log antes de deletar
        log = LogSistema(
            usuario_id=current_user.id,
            acao='REPORT_DELETED',
            detalhes={
                'relatorio_id': relatorio_id,
                'title': title
            }
        )
        db.session.add(log)
        
        db.session.delete(relatorio)
        db.session.commit()
        
        return jsonify({'message': 'Relatório deletado com sucesso'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Erro interno: {str(e)}'}), 500
