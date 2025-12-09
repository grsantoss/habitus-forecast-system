import requests
import io

BASE_URL = "http://localhost:5000"
LOGIN_URL = f"{BASE_URL}/api/auth/login"
UPLOAD_URL = f"{BASE_URL}/api/upload-planilha"
TIMEOUT = 30

def test_post_api_upload_planilha():
    # Authenticate to get JWT token
    auth_payload = {
        "email": "admin@habitus.com",
        "password": "admin123"
    }
    login_resp = requests.post(LOGIN_URL, json=auth_payload, timeout=TIMEOUT)
    assert login_resp.status_code == 200, f"Login failed: {login_resp.text}"
    login_data = login_resp.json()
    token = login_data.get("access_token") or login_data.get("token") or login_data.get("jwt")
    assert token, "JWT token not found in login response"

    headers = {
        "Authorization": f"Bearer {token}"
    }

    # Prepare a simple valid Excel file content in memory
    # Excel file minimal content (xlsx) header bytes for empty file could be tricky, so make a small valid xlsx content.
    # Here we create a simple Excel file using openpyxl for the upload without writing to disk.

    try:
        import openpyxl
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws.append([123, 456])
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
    except ImportError:
        # fallback dummy content if openpyxl not installed
        # This won't be a valid xlsx but the test expects valid Excel spreadsheet so better raise error to install openpyxl.
        raise RuntimeError("openpyxl is required for this test to create valid Excel file")

    files = {
        "file": ("test_planilha.xlsx", excel_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    }

    resp = requests.post(UPLOAD_URL, headers=headers, files=files, timeout=TIMEOUT)
    assert resp.status_code == 200, f"Upload failed: {resp.status_code} {resp.text}"

    resp_json = resp.json()
    # Basic validation on response assuming JSON with success or processed info
    assert isinstance(resp_json, dict), "Response is not JSON dictionary"
    assert "message" in resp_json or "success" in resp_json or resp_json.get("status") == "ok", f"Unexpected response content: {resp_json}"

test_post_api_upload_planilha()